import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

class TemplateParser:
    def __init__(self, file_buffer):
        try:
            self.wb = openpyxl.load_workbook(file_buffer, data_only=True)
            self.valid = True
        except:
            self.valid = False
            self.wb = None
        self.detected_theme = None

    def _argb_to_hex(self, argb):
        """Converts Excel ARGB (FF003366) to Hex (#003366). Returns None if invalid."""
        if not argb or not isinstance(argb, str): return None
        if len(argb) == 8: return "#" + argb[2:]
        return "#" + argb

    def parse(self):
        """
        Scans workbook and returns a structured 'Blueprint' of components.
        """
        blueprint = []
        
        if not self.valid or not self.wb:
            return blueprint

        # 1. Global Theme Guess (Scan first sheet, first few rows)
        try:
            first_sheet = self.wb.worksheets[0]
            color_counts = {}
            for row in first_sheet.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    if cell.fill and cell.fill.start_color and cell.fill.start_color.type == 'rgb':
                        c = self._argb_to_hex(cell.fill.start_color.rgb)
                        if c and c != '#FFFFFF' and c != '#000000':
                            color_counts[c] = color_counts.get(c, 0) + 1
            self.detected_theme = max(color_counts, key=color_counts.get) if color_counts else '#003366'
        except:
            self.detected_theme = '#003366'

        # 2. Iterate Sheets
        for sheet in self.wb.worksheets:
            # --- TOC DETECTION ---
            # Check for common names for Table of Contents
            is_toc = False
            clean_name = sheet.title.strip().lower()
            if clean_name in ['table of contents', 'contents', 'toc', 'index', 'agenda']:
                is_toc = True
            
            sheet_map = {
                'sheet_name': sheet.title,
                'is_toc': is_toc,
                'components': self._scan_sheet(sheet)
            }
            blueprint.append(sheet_map)
            
        return blueprint

    def _scan_sheet(self, sheet):
        components = []
        visited = set()
        
        # Limit scan to reasonable bounds to prevent hanging on massive empty sheets
        max_row = min(sheet.max_row, 1000) 
        max_col = min(sheet.max_column, 50)

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell_ref = (row_idx, col_idx)
                if cell_ref in visited: continue
                
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value is None: continue

                # Found unvisited content -> Trace the block
                block = self._trace_block(sheet, row_idx, col_idx, max_row, max_col)
                
                # Mark as visited
                for r in range(block['min_r'], block['max_r'] + 1):
                    for c in range(block['min_c'], block['max_c'] + 1):
                        visited.add((r, c))

                # Analyze the block
                comp = self._analyze_block(sheet, block)
                if comp: components.append(comp)

        return components

    def _trace_block(self, sheet, start_r, start_c, limit_r, limit_c):
        """Rectangular expansion to find contiguous data."""
        curr_r, curr_c = start_r, start_c
        
        # Expand Right
        while curr_c < limit_c and sheet.cell(row=start_r, column=curr_c + 1).value is not None:
            curr_c += 1
        max_c = curr_c
        
        # Expand Down (checking full width)
        while curr_r < limit_r:
            next_row_empty = False
            for c in range(start_c, max_c + 1):
                if sheet.cell(row=curr_r + 1, column=c).value is None:
                    next_row_empty = True
                    break
            if next_row_empty: break
            curr_r += 1
        max_r = curr_r
        
        return {'min_r': start_r, 'max_r': max_r, 'min_c': start_c, 'max_c': max_c}

    def _analyze_block(self, sheet, block):
        min_r, max_r, min_c, max_c = block['min_r'], block['max_r'], block['min_c'], block['max_c']
        
        width = max_c - min_c + 1
        height = max_r - min_r + 1
        start_cell = sheet.cell(row=min_r, column=min_c)
        
        # HEURISTIC 1: Single Cell or Small 1-Row Block -> TITLE / TEXT
        if height == 1 and width < 3:
            style = {}
            if start_cell.font and start_cell.font.color and start_cell.font.color.type == 'rgb':
                 style['font_color'] = self._argb_to_hex(start_cell.font.color.rgb)
            if start_cell.fill and start_cell.fill.start_color.type == 'rgb':
                 style['bg_color'] = self._argb_to_hex(start_cell.fill.start_color.rgb)
            
            return {
                'type': 'text',
                'value': start_cell.value,
                'row': min_r,
                'col': min_c,
                'style': style
            }

        # HEURISTIC 2: Larger Block -> DATAFRAME
        if height > 1:
            headers = []
            for c in range(min_c, max_c + 1):
                val = sheet.cell(row=min_r, column=c).value
                headers.append(str(val) if val else f"Col_{c}")

            # Extract Header Style for overrides
            header_style = {}
            if start_cell.fill and start_cell.fill.start_color.type == 'rgb':
                header_style['header_bg'] = self._argb_to_hex(start_cell.fill.start_color.rgb)
            if start_cell.font and start_cell.font.color and start_cell.font.color.type == 'rgb':
                header_style['header_font'] = self._argb_to_hex(start_cell.font.color.rgb)
            
            return {
                'type': 'dataframe',
                'headers': headers,
                'rows': height,
                'row': min_r,
                'col': min_c,
                'style': header_style
            }
        
        return None