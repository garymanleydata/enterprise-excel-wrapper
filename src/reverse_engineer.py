import openpyxl
from openpyxl.utils import get_column_letter
import os
import re

class EnterpriseExcelDecompiler:
    """
    Reverse engineers an existing Excel file into a Python script.
    v3.4: Enhanced font detection and merge parameter integration for fAddText.
    """
    def __init__(self, vInputPath, vHints=None):
        self.vInputPath = vInputPath
        # Support file path or bytes buffer
        self.vWorkbook = openpyxl.load_workbook(vInputPath, data_only=True)
        self.vHints = vHints or {}
        
        self.vGlobalStartCol = self.vHints.get('GlobalStartCol', 1)
        self.vIgnoredSheets = self.vHints.get('IgnoredSheets', [])
        
        self.vCodeLines = []
        self.vImports = [
            "import pandas as pd",
            "import sys",
            "import os",
            "# Add src to path",
            "sys.path.append(os.path.abspath('src'))",
            "from enterprise_writer import EnterpriseExcelWriter"
        ]

    def _fGetHexColor(self, vColorObj):
        """Extracts hex color from openpyxl color object."""
        try:
            if not vColorObj or not hasattr(vColorObj, 'rgb'): return None
            vHex = vColorObj.rgb
            if isinstance(vHex, str) and vHex != '00000000':
                return "#" + vHex[2:] if len(vHex) > 6 else "#" + vHex
            return None
        except:
            return None

    def _fCleanString(self, vStr):
        if vStr is None: return ""
        return str(vStr).replace("'", "\\'").replace('"', '\\"')

    def fExtractTheme(self):
        """Identifies primary color from the first non-ignored sheet."""
        for name in self.vWorkbook.sheetnames:
            if name in self.vIgnoredSheets: continue
            sheet = self.vWorkbook[name]
            colors = {}
            for row in sheet.iter_rows(max_row=20):
                for cell in row:
                    c = self._fGetHexColor(cell.fill.start_color)
                    if c and c not in ['#FFFFFF', '#000000']:
                        colors[c] = colors.get(c, 0) + 1
            if colors:
                return max(colors, key=colors.get)
        return "#003366"

    def fScanSheet(self, vSheet, vSheetName):
        self.vCodeLines.append(f"\n# --- Sheet: {vSheetName} ---")
        self.vCodeLines.append(f"vReport.fNewSheet('{vSheetName}')")
        
        # 1. Capture Column Widths
        vWidths = {}
        for i in range(1, 20): # Check first 20 columns
            vColLet = get_column_letter(i)
            if vColLet in vSheet.column_dimensions:
                w = vSheet.column_dimensions[vColLet].width
                if w and w != 8.43: 
                    vWidths[i] = round(w, 1)
        
        if vWidths:
            self.vCodeLines.append(f"vReport.fSetColumnWidths({vWidths})")

        # 2. Gridlines
        if hasattr(vSheet, 'sheet_view') and vSheet.sheet_view and not vSheet.sheet_view.showGridLines:
            self.vCodeLines.append("# Note: Gridlines hidden in source")

        vSheetHints = self.vHints.get('Sheets', {}).get(vSheetName, {}).get('Components', {})
        vMaxRow = vSheet.max_row
        vCurrentRow = 1
        vSkipCounter = 0
        
        while vCurrentRow <= vMaxRow:
            vRowKey = str(vCurrentRow)
            
            # A. Process User-Defined Hints (Dataframes)
            if vRowKey in vSheetHints:
                if vSkipCounter > 0:
                    self.vCodeLines.append(f"vReport.fSkipRows({vSkipCounter})")
                    vSkipCounter = 0
                
                vHint = vSheetHints[vRowKey]
                vVar = vHint.get('var_name', f"df_{vSheetName.replace(' ', '_')}")
                vTot = vHint.get('add_totals', False)
                vFilt = vHint.get('auto_filter', True)
                self.vCodeLines.append(f"vReport.fWriteDataframe({vVar}, vStartCol={self.vGlobalStartCol}, vAddTotals={vTot}, vAutoFilter={vFilt})")
                vCurrentRow += vHint.get('skip_rows', 1)
                continue

            # B. Check for values across the scanning range (Column A onwards)
            vTargetCell = None
            vMergedRange = None
            
            # Scan columns in this row to find the first content
            for c_idx in range(1, self.vGlobalStartCol + 2):
                vCell = vSheet.cell(row=vCurrentRow, column=c_idx)
                
                vFoundMerge = None
                for rng in vSheet.merged_cells.ranges:
                    if rng.min_row == vCurrentRow and rng.min_col == c_idx:
                        vFoundMerge = rng
                        break
                
                if vCell.value is not None or vFoundMerge:
                    vTargetCell = vCell
                    vMergedRange = vFoundMerge
                    break

            if not vTargetCell:
                vSkipCounter += 1
                vCurrentRow += 1
                continue
            
            # Flush skip buffer
            if vSkipCounter > 0:
                self.vCodeLines.append(f"vReport.fSkipRows({vSkipCounter})")
                vSkipCounter = 0

            # Style Extraction
            vBg = self._fGetHexColor(vTargetCell.fill.start_color)
            vFg = self._fGetHexColor(vTargetCell.font.color)
            vBld = vTargetCell.font.b
            vSize = vTargetCell.font.size
            vFontName = vTargetCell.font.name
            
            vParams = []
            if vBg: vParams.append(f"vBgColour='{vBg}'")
            if vFg: vParams.append(f"vFontColour='{vFg}'")
            if vBld: vParams.append("vBold=True")
            if vSize and vSize != 10: vParams.append(f"vFontSize={vSize}")
            if vFontName and vFontName != "Arial": vParams.append(f"vFontName='{vFontName}'")
            
            # Calculate column offset relative to GlobalStartCol
            vColOffset = vTargetCell.column - 1
            if vColOffset != self.vGlobalStartCol:
                vParams.append(f"vStartCol={vColOffset}")
            
            # Detect Merge
            if vMergedRange:
                vParams.append("vMerge=True")
            
            vParamStr = ", ".join(vParams)
            if vParamStr: vParamStr = ", " + vParamStr
            
            vCleanVal = self._fCleanString(vTargetCell.value)

            # Heuristic Decision: Use fAddTitle for large merged banners, otherwise fAddText
            if vMergedRange and vSize and vSize >= 14:
                self.vCodeLines.append(f"vReport.fAddTitle('{vCleanVal}'{vParamStr})")
                # Move to the end of the merge height to avoid double processing
                vHeight = vMergedRange.max_row - vMergedRange.min_row
                vCurrentRow += vHeight
            else:
                self.vCodeLines.append(f"vReport.fAddText('{vCleanVal}'{vParamStr})")

            vCurrentRow += 1

    def fGenerateCode(self):
        vTheme = self.fExtractTheme()
        vFullCode = "\n".join(self.vImports) + "\n\n"
        vFullCode += f"# 1. Configuration\n"
        vFullCode += f"vConfig = {{'Global': {{'primary_colour': '{vTheme}', 'hide_gridlines': 'True'}}, 'Header': {{'bg_colour': '{vTheme}', 'font_size': 12}}}}\n\n"
        vFullCode += "# 2. Report Generation\n"
        vFullCode += f"vReport = EnterpriseExcelWriter('Recreated_Report.xlsx', vConfig=vConfig, vGlobalStartCol={self.vGlobalStartCol})\n"
        
        self.vCodeLines = []
        for name in self.vWorkbook.sheetnames:
            if name in self.vIgnoredSheets: continue
            self.fScanSheet(self.vWorkbook[name], name)
            
        vFullCode += "\n".join(self.vCodeLines)
        
        if self.vHints.get('GenerateTOC', True):
            vFullCode += "\n\nvReport.fGenerateTOC()"
            
        vFullCode += "\n\nvReport.fClose()\n"
        return vFullCode